import numpy as np
# import xlwt
# import xlutils.copy
# import xlrd
import os, yaml, re
from openpyxl import Workbook, load_workbook
from tqdm import tqdm


def kappa(matrix):
    n = np.sum(matrix)
    sum_po = 0
    sum_pe = 0
    for i in range(len(matrix[0])):
        sum_po += matrix[i][i]
        row = np.sum(matrix[i, :])
        col = np.sum(matrix[:, i])
        sum_pe += row * col
    po = sum_po / n
    pe = sum_pe / (n * n)
    # print(po, pe)
    return (po - pe) / (1 - pe)


def aa_oa_(matrix, expo, mname, t1, t2, time=0):  # number表示同时打印几个
    number = 0
    savepath = expo + "_" + mname + '_' + str(number) + "_matrix.npy"
    while os.path.exists(savepath):
        number += 1
        savepath = expo + "_" + mname + '_' + str(number) + "_matrix.npy"
    np.save(savepath, matrix)
    accuracy = []
    b = np.sum(matrix, axis=0)
    c = 0
    book = xlwt.Workbook(encoding='utf-8', style_compression=0)
    sheet = book.add_sheet('实验结果', cell_overwrite_ok=True)
    sheet.write(0, 0, "Category")
    sheet.write(1, 0, "Overall")
    sheet.write(2, 0, "Correct")
    sheet.write(3, 0, "Accuracy")
    for i in range(1, matrix.shape[0]):
        sheet.write(0, i, i)
        a = matrix[i][i]/b[i]
        c += matrix[i][i]
        accuracy.append(a)
        sheet.write(1, i, b[i])
        sheet.write(2, i, matrix[i][i])
        sheet.write(3, i, a)
        tqdm.write("Category:{}. Overall:{}. Correct:{}. Accuracy:{:.6f}".format(i, b[i], matrix[i][i], a))
    aa = np.mean(accuracy)
    oa = c / np.sum(b, axis=0)
    k = kappa(matrix)
    tqdm.write("OA:{:.6f} AA:{:.6f} Kappa:{:.6f}".format(oa, aa, k))
    sheet.write(5, 1, "OA")
    sheet.write(5, 2, oa)
    sheet.write(5, 3, "AA")
    sheet.write(5, 4, aa)
    sheet.write(5, 5, "KAPPA")
    sheet.write(5, 6, k)
    sheet.write(5, 7, "Train time(s)")
    sheet.write(5, 8, t1)
    sheet.write(5, 9, "Test time(s)")
    sheet.write(5, 10, t2)
    savepath = expo + "_" + mname + "_" + str(number) + '_result.xls'
    book.save(savepath)
    return number


def aa_oa(matrix):
    accuracy = []
    b = np.sum(matrix, axis=0)
    c = 0
    on_display = []
    for i in range(1, matrix.shape[0]):
        a = matrix[i][i]/b[i]
        c += matrix[i][i]
        accuracy.append(a)
        on_display.append([b[i], matrix[i][i], a])
        print("Category:{}. Overall:{}. Correct:{}. Accuracy:{:.6f}".format(i, b[i], matrix[i][i], a))
    aa = np.mean(accuracy)
    oa = c / np.sum(b, axis=0)
    k = kappa(matrix)
    print("OA:{:.6f} AA:{:.6f} Kappa:{:.6f}".format(oa, aa, k))
    return [aa, oa, k, on_display]


def expo_result(result, cfg, time, group_num):
    savepath = cfg['RESULT_excel']
    col = group_num * 8
    if group_num == 0:
        wb = Workbook()
    else:
        wb = load_workbook(savepath)
    sheet = wb.active
    sheet.cell(1+col, 1, "Category")
    sheet.cell(2+col, 1, "Overall")
    sheet.cell(3+col, 1, "Correct")
    sheet.cell(4+col, 1, "Accuracy")
    for i in range(len(result[3])):
        sheet.cell(1+col, i+2, i+1)
        sheet.cell(2+col, i+2, result[3][i][0])
        sheet.cell(3+col, i+2, result[3][i][1])
        sheet.cell(4+col, i+2, result[3][i][2])
    sheet.cell(6+col, 2, "OA")
    sheet.cell(6+col, 3, result[1])
    sheet.cell(6+col, 4, "AA")
    sheet.cell(6+col, 5, result[0])
    sheet.cell(6+col, 6, "KAPPA")
    sheet.cell(6+col, 7, result[2])
    sheet.cell(6+col, 8, "Train time(s)")
    sheet.cell(6+col, 9, time[0])
    sheet.cell(6+col, 10, "Test time(s)")
    sheet.cell(6 + col, 11, time[0])
    if group_num == 0:
        config_sheet = wb.create_sheet(title='config')
        config_sheet.row_dimensions[1].width = 20
        print_dict_data(cfg, config_sheet, 1, 1)
    wb.save(savepath)

    # col = group_num * 8  # 每一组占表格8格
    # if group_num == 0:
    #     book = xlwt.Workbook(encoding='utf-8', style_compression=0)
    #     sheet = book.add_sheet('实验结果', cell_overwrite_ok=True)
    # else:
    #     book = xlrd.open_workbook(savepath, formatting_info=True)
    #     book = xlutils.copy.copy(book)
    #     sheet = book.get_sheet(0)

    # 如需保存矩阵，可以将下方代码复用
def print_dict_data(data, sheet, row, column):
    for key, value in data.items():
        if key != 'DATA_DICT':
            sheet.cell(row=row, column=column, value=key)
            column += 1
            if isinstance(value, dict):
                row, column = print_dict_data(value, sheet, row, column)
                column -= 1
            elif isinstance(value, list):
                row = print_list_data(value, sheet, row, column)
                column -= 1
            else:
                sheet.cell(row=row, column=column, value=value)
                row += 1
                column -= 1
    return row, column


def print_list_data(data, sheet, row, column):
    for item in data:
        if isinstance(item, dict):
            row, column = print_dict_data(item, sheet, row, column)
        elif isinstance(item, list):
            row = print_list_data(item, sheet, row, column)
        else:
            sheet.cell(row=row, column=column, value=item)
            column += 1

    return row + 1



class yml2Excel(object):

    def __init__(self):
        self.file_path_list = []
        self.save_excel_list = []

    # 获取路径下所有yml文件
    def print_all_path(self, init_file_path, keyword):
        for cur_dir, sub_dir, include_file in os.walk(init_file_path):
            if include_file:
                for file in include_file:
                    if re.search(keyword, file):
                        self.file_path_list.append(cur_dir + file)
        return self.file_path_list

    # 遍历每个yml文件
    def open_yml_file(self, savepath):
        # 打开excel 
        book = xlwt.Workbook(encoding = 'utf-8')
        # 把excel中sheet的名字变成动态的
        sheet = locals()

        # 统计第几个yml文件,一个文件信息保存一行,用来作为excel中的行数
        count = 0
        for yml_file in self.file_path_list:
            if (count%2 == 0):
                count = 0
            count  = count +1
            # 获取文件名前缀来命名excel文件
            cut_filename = yml_file.split("\\")[-1]
            # 这里设置每一个sheet的名字 
            if ((count+1)%2 == 0):
                sheet[cut_filename[:-4]] = book.add_sheet(cut_filename[:-4], cell_overwrite_ok=True)
                upname = sheet[cut_filename[:-4]]
            else:
                sheet[cut_filename[:-4]] = upname
            # 保存最终需要的数据信息
            all_data = []
            # 统计是第几个数组
            count_matirx = 0

            path_yml = "%s" % yml_file
            # print(path_yml)
            with open(path_yml,"r",encoding='utf-8') as f:
                yml = f.readlines()
                col = 0
                for line in yml:
                    sheet[cut_filename[:-4]].write(col, 0, line)
                    # sheet.write(col, 1, line[1])
                    col += 1
        # 保存数据的excel文件路径
        book.save(savepath)



def indicator(test_matrix, result_file, consume_time, group_num):
    result = aa_oa(test_matrix)
    expo_result(result, result_file, consume_time, group_num)


def indicator_SCLN(test_matrix):
    for i in range(4):
        aa, oa, correct, k = aa_oa(test_matrix[i], i)
        print("Kappa:{:.6f} AA:{:.6f} OA:{:.6f}\n".format(k, aa, oa))

